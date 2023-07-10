from openpyxl.styles import PatternFill
from datetime import datetime, timedelta

def cell_colour(basic_value: int, value: int, cell) -> None:
    green_fill_cell = PatternFill(patternType='solid', 
                           fgColor='35FC03')
    red_fill_cell = PatternFill(patternType='solid', 
                           fgColor='FC2C03')
    if value > basic_value:
        cell.fill = green_fill_cell
    elif value < basic_value:
        cell.fill = red_fill_cell
    else:
        pass

def days(n = 12) -> list:
    now = datetime.now()
    time = now.time()
    update_currency_report_time = time.replace(hour=12, minute=15, second=0, microsecond=0)
    if time > update_currency_report_time:
        last_report_day = datetime.today()
    else: 
        last_report_day = datetime.today() - timedelta(days=1)  
    days = [datetime.strftime(last_report_day - timedelta(days=i), "%Y-%m-%d") for i in range(n)]
    days.reverse()
    return days

def is_weekend(dateStr: str) -> bool:
    dateObj = datetime.strptime(dateStr, '%Y-%m-%d').date()
    return dateObj.weekday() >= 5

