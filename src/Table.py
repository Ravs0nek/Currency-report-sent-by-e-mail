from Api import CurrencyData as cd
from datetime import datetime
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl import Workbook
from SecondaryFunctions import cell_colour, days

class ExcelFile:
    
    def __init__(self):
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        

    def title(self,title):
        self.sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=18)
        cell = self.sheet.cell(row=1, column=1)
        cell.alignment = Alignment(horizontal="center")
        cell.value = title
        cell.font = Font(size=20, bold = True)
        for i in range(18):
            cell = self.sheet.cell(row = 1, column = 1 + i)
            cell.border = Border(left=Side(style='thick'), 
                     right=Side(style='thick'), 
                     top=Side(style='thick'), 
                     bottom=Side(style='thick'))
        self.workbook.save(filename = "CurrencyData.xlsx")

    def currency_report(self, currency, starting_row, starting_column):
        currency_data = cd.week_exchange_rate(exchange_rates=cd.exchange_rate_api(currency=currency))
        rates, dates = currency_data[0], currency_data[1]
        basic_value = rates[0]
        self.sheet.merge_cells(start_row=starting_row, start_column=starting_column, 
                               end_row=starting_row, end_column=starting_column + 7)
        title_cell = self.sheet.cell(row=starting_row, column=starting_column)
        title_cell.value = f'Exchange rate {currency}/PLN'
        title_cell.font = Font(size = 13, bold=True)
        name_date_cell = self.sheet.cell(row=starting_row + 1, column=starting_column)
        name_date_cell.value = 'Dates:'
        name_rate_cell = self.sheet.cell(row=starting_row + 2, column=starting_column)
        name_rate_cell.value = 'Rates:'
        name_rate_cell.font = name_date_cell.font = Font(bold=True)
        name_rate_cell.alignment = name_date_cell.alignment = title_cell.alignment = Alignment(horizontal="center")
        for i in range(7): #week length
            date_cell = self.sheet.cell(row=starting_row + 1, column=starting_column + (i + 1))
            date_cell.value = dates[i]
            rate_cell = self.sheet.cell(row=starting_row + 2, column=starting_column + (i + 1))
            rate_cell.value = rates[i]
            date_cell.alignment = rate_cell.alignment = Alignment(horizontal="center") 
            cell_colour(basic_value = basic_value, value = rate_cell.value, cell = rate_cell)
        self.workbook.save(filename = "CurrencyData.xlsx")

